import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
import datetime
from docx import Document
import os
import subprocess
import platform

# File paths
EXCEL_FILE = "vehicle_data.xlsx"
BILLS_FOLDER = "bills"
TEMPLATE_PATH = "template.docx"

# Ensure bills folder exists
if not os.path.exists(BILLS_FOLDER):
    os.makedirs(BILLS_FOLDER)

# Function to load the current load sheet counter
def load_counter():
    if os.path.exists("counter.txt"):
        with open("counter.txt", "r") as file:
            return int(file.read())
    else:
        return 1  # Starting from GM0001 if the file doesn't exist

# Function to save the current load sheet counter
def save_counter(counter):
    with open("counter.txt", "w") as file:
        file.write(str(counter))

# Function to handle form submission
def submit_form():
    global load_sheet_counter

    # Get user inputs
    vehicle_no = entry_vehicle_no.get().strip()
    tare_wt = entry_tare_wt.get().strip()
    gross_wt = entry_gross_wt.get().strip()

    # Validate inputs
    if not vehicle_no:
        messagebox.showerror("Error", "Vehicle number cannot be empty!")
        return

    try:
        tare_wt = float(tare_wt)
        gross_wt = float(gross_wt)
        if tare_wt < 0 or gross_wt < 0:
            raise ValueError("Weights cannot be negative.")
    except ValueError as e:
        messagebox.showerror("Error", f"Invalid weight values: {e}")
        return

    # Calculate Net Weight (ensure it's always positive)
    net_wt = abs(gross_wt - tare_wt)

    # Generate Load Sheet Number (GM0001, GM0002, etc.)
    load_sheet_number = f"GM{load_sheet_counter:04d}"
    load_sheet_counter += 1
    save_counter(load_sheet_counter)  # Save the updated counter

    # Get current date
    current_date = datetime.datetime.now().strftime("%m/%d/%Y %I:%M:%S %p")

    # Save data to Excel
    save_to_excel(load_sheet_number, current_date, vehicle_no, tare_wt, gross_wt, net_wt)

    # Add record to the interface
    add_record_to_interface(load_sheet_number, current_date, vehicle_no, tare_wt, gross_wt, net_wt)

    # Show success message
    messagebox.showinfo("Success", "Bill generated and data saved!")
    update_status("Bill generated and data saved!")

    # Generate Word bill based on template
    output_docx = generate_word_bill(load_sheet_number, current_date, vehicle_no, tare_wt, gross_wt, net_wt)

    # Open the generated Word document
    if os.path.exists(output_docx):
        try:
            if platform.system() == "Windows":
                os.startfile(output_docx)  # Open the document on Windows
            elif platform.system() == "Darwin":  # macOS
                subprocess.run(["open", output_docx])  # Open the document on macOS
            else:
                messagebox.showwarning("Unsupported OS", "Automatic file opening is not supported on this operating system.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open the document: {e}")
    else:
        messagebox.showerror("Error", "Failed to generate the bill!")

    # Clear form fields
    clear_form()

# Function to save data to Excel
def save_to_excel(load_sheet_number, date, vehicle_no, tare_wt, gross_wt, net_wt):
    try:
        # Load or create Excel file
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            # Add headers if the file is new
            ws.append(["Load Sheet Number", "Date", "Vehicle No", "Tare Wt (Kgs)", "Gross Wt (Kgs)", "Net Wt (Kgs)"])

        # Append new row
        ws.append([load_sheet_number, date, vehicle_no, tare_wt, gross_wt, net_wt])

        # Save the file
        wb.save(EXCEL_FILE)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save data: {e}")

# Function to add record to the interface
def add_record_to_interface(load_sheet_number, date, vehicle_no, tare_wt, gross_wt, net_wt):
    record = [load_sheet_number, date, vehicle_no, tare_wt, gross_wt, net_wt]
    tree.insert("", tk.END, values=record)

# Function to load existing records from Excel
def load_existing_records():
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        # Clear existing records in the Treeview
        for row in tree.get_children():
            tree.delete(row)
        # Load records from Excel
        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
            if row:  # Check if row is not empty
                tree.insert("", tk.END, values=row)
    except FileNotFoundError:
        pass  # No existing file, skip loading

# Function to generate bill based on Word template
def generate_word_bill(load_sheet_number, date, vehicle_no, tare_wt, gross_wt, net_wt):
    bill_data = {
        "LOAD_SHEET_NO": load_sheet_number,
        "DATE": date,
        "VEHICLE_NO": vehicle_no,
        "TARE_WEIGHT": tare_wt,
        "GROSS_WEIGHT": gross_wt,
        "NET_WEIGHT": net_wt
    }

    # Load Word template
    doc = Document(TEMPLATE_PATH)

    # Replace placeholders in the template
    for paragraph in doc.paragraphs:
        for key, value in bill_data.items():
            paragraph.text = paragraph.text.replace(f"{{{{{key}}}}}", str(value))

    # Save the filled document
    output_docx = f"{BILLS_FOLDER}/Bill_{bill_data['LOAD_SHEET_NO']}.docx"
    doc.save(output_docx)

    messagebox.showinfo("Success", f"Bill saved as Word document:\n{output_docx}")
    return output_docx  # Return the path of the generated document

# Clear form fields
def clear_form():
    entry_vehicle_no.delete(0, tk.END)
    entry_tare_wt.delete(0, tk.END)
    entry_gross_wt.delete(0, tk.END)
    update_status("Ready")

# Search functionality
def search_record():
    search_term = entry_search.get().strip()
    if not search_term:
        messagebox.showwarning("Search", "Please enter a search term!")
        return

    # Clear existing search results
    for row in tree.get_children():
        tree.delete(row)

    # Search in Excel file
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
            if search_term.lower() in str(row[2]).lower() or search_term.lower() in str(row[0]).lower():  # Vehicle No or Load Sheet No
                tree.insert("", tk.END, values=row)
    except FileNotFoundError:
        messagebox.showerror("Error", "No data found!")

# Delete functionality
def delete_record():
    selected_item = tree.selection()
    if not selected_item:
        messagebox.showwarning("Delete", "Please select a record to delete!")
        return

    confirm = messagebox.askyesno("Confirm", "Are you sure you want to delete this record?")
    if confirm:
        load_sheet_number = tree.item(selected_item, "values")[0]  # Get Load Sheet Number
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

            # Find the row to delete
            row_to_delete = None
            for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), start=2):  # Start from row 2
                if row[0] == load_sheet_number:
                    row_to_delete = idx
                    break

            if row_to_delete:
                ws.delete_rows(row_to_delete)  # Delete the row
                wb.save(EXCEL_FILE)
                tree.delete(selected_item)  # Remove from the table
                messagebox.showinfo("Success", "Record deleted successfully!")
                update_status("Record deleted successfully!")
            else:
                messagebox.showerror("Error", "Record not found in the Excel file!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete record: {e}")

# Refresh functionality
def refresh_table():
    load_existing_records()
    update_status("Table refreshed!")

# Help menu
def show_help():
    help_text = """
    **User Guide:**
    1. Enter the vehicle number, tare weight, and gross weight.
    2. Click 'Generate Bill' to save the data and create a bill.
    3. Use the 'Search' field to find records by vehicle number or load sheet number.
    4. Use the 'Delete' button to remove a record.
    5. Click 'Refresh' to update the table with the latest data.
    """
    messagebox.showinfo("Help", help_text)

# Update status bar
def update_status(message):
    status_bar.config(text=message)

# Initialize the main window
root = tk.Tk()
root.title("Vehicle Weight Bill Generator")

# Static text (company info)
company_info = """
कम्पनीको नाम: तपाईंको कम्पनी
PAN नम्बर: 123456789
फोन: +९७७-१२३४५६७८९०
स्थान: काठमाडौं, नेपाल
"""
label_company_info = tk.Label(root, text=company_info, justify=tk.LEFT)
label_company_info.grid(row=0, column=0, columnspan=3, padx=10, pady=10)

# Form fields
label_vehicle_no = tk.Label(root, text="वाहन नम्बर:")
label_vehicle_no.grid(row=1, column=0, padx=10, pady=5)
entry_vehicle_no = tk.Entry(root)
entry_vehicle_no.grid(row=1, column=1, padx=10, pady=5)

label_tare_wt = tk.Label(root, text="टार वजन (किलो):")
label_tare_wt.grid(row=2, column=0, padx=10, pady=5)
entry_tare_wt = tk.Entry(root)
entry_tare_wt.grid(row=2, column=1, padx=10, pady=5)

label_gross_wt = tk.Label(root, text="ग्रस वजन (किलो):")
label_gross_wt.grid(row=3, column=0, padx=10, pady=5)
entry_gross_wt = tk.Entry(root)
entry_gross_wt.grid(row=3, column=1, padx=10, pady=5)

# Submit button
submit_button = tk.Button(root, text="Generate Bill", command=submit_form)
submit_button.grid(row=4, column=0, padx=10, pady=10)

# Clear button
clear_button = tk.Button(root, text="Clear", command=clear_form)
clear_button.grid(row=4, column=1, padx=10, pady=10)

# Search field and button
label_search = tk.Label(root, text="Search by Vehicle No or Load Sheet No:")
label_search.grid(row=5, column=0, padx=10, pady=5)
entry_search = tk.Entry(root)
entry_search.grid(row=5, column=1, padx=10, pady=5)
search_button = tk.Button(root, text="Search", command=search_record)
search_button.grid(row=5, column=2, padx=10, pady=5)

# Treeview to display records
columns = ("Load Sheet Number", "Date", "Vehicle No", "Tare Wt (Kgs)", "Gross Wt (Kgs)", "Net Wt (Kgs)")
tree = ttk.Treeview(root, columns=columns, show="headings")
tree.grid(row=6, column=0, columnspan=3, padx=10, pady=10)

# Add scrollbars to the table
scroll_y = ttk.Scrollbar(root, orient=tk.VERTICAL, command=tree.yview)
scroll_y.grid(row=6, column=3, sticky="ns")
tree.configure(yscrollcommand=scroll_y.set)

scroll_x = ttk.Scrollbar(root, orient=tk.HORIZONTAL, command=tree.xview)
scroll_x.grid(row=7, column=0, columnspan=3, sticky="ew")
tree.configure(xscrollcommand=scroll_x.set)

# Define headings
for col in columns:
    tree.heading(col, text=col)

# Load existing records on startup
load_existing_records()

# Delete button
delete_button = tk.Button(root, text="Delete Record", command=delete_record)
delete_button.grid(row=8, column=0, padx=10, pady=10)

# Refresh button
refresh_button = tk.Button(root, text="Refresh", command=refresh_table)
refresh_button.grid(row=8, column=1, padx=10, pady=10)

# Status bar
status_bar = tk.Label(root, text="Ready", bd=1, relief=tk.SUNKEN, anchor=tk.W)
status_bar.grid(row=9, column=0, columnspan=3, sticky="ew")

# Help menu
menu_bar = tk.Menu(root)
root.config(menu=menu_bar)

help_menu = tk.Menu(menu_bar, tearoff=0)
help_menu.add_command(label="Help", command=show_help)
menu_bar.add_cascade(label="Help", menu=help_menu)

# Keyboard shortcuts
root.bind("<Control-s>", lambda event: submit_form())
root.bind("<Control-c>", lambda event: clear_form())

# Initialize load sheet counter
load_sheet_counter = load_counter()

# Run the main application
root.mainloop()